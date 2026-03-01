"""
Excel Workbook Parser for Lansport Analytics ETL
=================================================
Reads the Lansport accounting workbook (.xlsx) and returns structured
Python dataclasses ready for database loading.

Verified against: Bookkeeping Lansport Main (1).xlsx

Worksheets parsed:
  "Chart of Accounts"  — columns: Account #, Account Name, Account Type, Normal Balance
  "General Ledger"     — columns: Date, Account #, Account Name, Description, Debit, Credit

Key notes:
  - Dates in the GL are stored as Excel serial numbers (days since 1899-12-30)
  - Section-header rows (e.g. "ASSETS", "LIABILITIES") in the CoA have no account number
    and are skipped automatically
  - Trailing empty rows in the GL are skipped
"""

import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)

# Excel epoch: serial 1 = 1900-01-01 (with the 1900 leap-year bug offset)
_EXCEL_EPOCH = datetime(1899, 12, 30)


# ─────────────────────────────────────────────────────────────────────────────
# Data structures
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class AccountRecord:
    account_number: str
    name: str
    type: str           # ASSET | LIABILITY | EQUITY | REVENUE | EXPENSE | CONTRA_ASSET
    normal_balance: str # DEBIT | CREDIT


@dataclass
class JournalEntryRecord:
    date: datetime
    account_number: str
    description: str
    debit: Decimal
    credit: Decimal


@dataclass
class WorkbookData:
    accounts: list[AccountRecord] = field(default_factory=list)
    journal_entries: list[JournalEntryRecord] = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
# Lookup tables  (lower-cased for comparison)
# ─────────────────────────────────────────────────────────────────────────────

ACCOUNT_TYPE_MAP: dict[str, str] = {
    "asset": "ASSET",
    "assets": "ASSET",
    "current asset": "ASSET",
    "non-current asset": "ASSET",
    "liability": "LIABILITY",
    "liabilities": "LIABILITY",
    "current liability": "LIABILITY",
    "equity": "EQUITY",
    "owner's equity": "EQUITY",
    "owners equity": "EQUITY",
    "capital": "EQUITY",
    "revenue": "REVENUE",
    "revenues": "REVENUE",
    "income": "REVENUE",
    "expense": "EXPENSE",
    "expenses": "EXPENSE",
    "cost": "EXPENSE",
    "contra asset": "CONTRA_ASSET",
    "contra_asset": "CONTRA_ASSET",
    "contraasset": "CONTRA_ASSET",
}

NORMAL_BALANCE: dict[str, str] = {
    "ASSET": "DEBIT",
    "LIABILITY": "CREDIT",
    "EQUITY": "CREDIT",
    "REVENUE": "CREDIT",
    "EXPENSE": "DEBIT",
    "CONTRA_ASSET": "CREDIT",
}

# Column header aliases — all lower-cased for matching
COA_NUMBER_ALIASES  = {"account #", "account number", "account no", "account#", "no", "number", "acct no"}
COA_NAME_ALIASES    = {"account name", "name", "description", "account description"}
COA_TYPE_ALIASES    = {"account type", "type", "category"}

GL_DATE_ALIASES     = {"date", "transaction date", "txn date"}
GL_ACCOUNT_ALIASES  = {"account #", "account number", "account no", "account#", "account", "acct no"}
GL_DESC_ALIASES     = {"description", "memo", "narration", "details", "particulars"}
GL_DEBIT_ALIASES    = {"debit", "dr", "debit amount"}
GL_CREDIT_ALIASES   = {"credit", "cr", "credit amount"}


# ─────────────────────────────────────────────────────────────────────────────
# Parser
# ─────────────────────────────────────────────────────────────────────────────

class ExcelParser:
    """Parses a Lansport accounting workbook and returns structured data."""

    CHART_SHEET  = "Chart of Accounts"
    LEDGER_SHEET = "General Ledger"

    def __init__(self, file_path: str):
        self.file_path = file_path

    def parse(self) -> WorkbookData:
        logger.info(f"Opening workbook: {self.file_path}")
        wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)
        try:
            data = WorkbookData()
            data.accounts = self._parse_chart_of_accounts(wb)
            data.journal_entries = self._parse_general_ledger(wb)
            return data
        finally:
            wb.close()

    # ── Chart of Accounts ─────────────────────────────────────────────────

    def _parse_chart_of_accounts(self, wb) -> list[AccountRecord]:
        sheet = self._require_sheet(wb, self.CHART_SHEET)
        accounts: list[AccountRecord] = []
        col_map: dict[str, int] = {}

        for row in sheet.iter_rows(values_only=True):
            if not any(c is not None for c in row):
                continue

            if not col_map:
                row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
                col_map = self._detect_columns(row_lower, {
                    "account_number": COA_NUMBER_ALIASES,
                    "name":           COA_NAME_ALIASES,
                    "type":           COA_TYPE_ALIASES,
                })
                if col_map:
                    logger.debug(f"CoA column map: {col_map}")
                continue

            acct_no  = self._str_cell(row, col_map, "account_number")
            name     = self._str_cell(row, col_map, "name")
            type_raw = self._str_cell(row, col_map, "type")

            # Skip section-header rows (e.g. "ASSETS", "LIABILITIES")
            if not acct_no or not name:
                continue

            acct_type = ACCOUNT_TYPE_MAP.get(type_raw.lower(), "EXPENSE")
            accounts.append(AccountRecord(
                account_number=acct_no,
                name=name,
                type=acct_type,
                normal_balance=NORMAL_BALANCE.get(acct_type, "DEBIT"),
            ))

        logger.info(f"Chart of Accounts: {len(accounts)} accounts parsed")
        return accounts

    # ── General Ledger ────────────────────────────────────────────────────

    def _parse_general_ledger(self, wb) -> list[JournalEntryRecord]:
        sheet = self._require_sheet(wb, self.LEDGER_SHEET)
        entries: list[JournalEntryRecord] = []
        col_map: dict[str, int] = {}

        for row in sheet.iter_rows(values_only=True):
            if not any(c is not None for c in row):
                continue

            if not col_map:
                row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
                col_map = self._detect_columns(row_lower, {
                    "date":           GL_DATE_ALIASES,
                    "account_number": GL_ACCOUNT_ALIASES,
                    "description":    GL_DESC_ALIASES,
                    "debit":          GL_DEBIT_ALIASES,
                    "credit":         GL_CREDIT_ALIASES,
                })
                if col_map:
                    logger.debug(f"GL column map: {col_map}")
                continue

            raw_date   = self._raw_cell(row, col_map, "date")
            acct_no    = self._str_cell(row, col_map, "account_number")
            description = self._str_cell(row, col_map, "description")
            debit      = self._decimal_cell(row, col_map, "debit")
            credit     = self._decimal_cell(row, col_map, "credit")

            parsed_date = self._parse_date(raw_date)

            # Skip rows with no valid date or account number (e.g. trailing empty rows)
            if not parsed_date or not acct_no:
                continue

            entries.append(JournalEntryRecord(
                date=parsed_date,
                account_number=acct_no,
                description=description,
                debit=debit,
                credit=credit,
            ))

        logger.info(f"General Ledger: {len(entries)} entries parsed")
        return entries

    # ── Helpers ───────────────────────────────────────────────────────────

    @staticmethod
    def _require_sheet(wb, name: str):
        if name not in wb.sheetnames:
            available = ", ".join(f'"{s}"' for s in wb.sheetnames)
            raise ValueError(f'Sheet "{name}" not found. Available: {available}')
        return wb[name]

    @staticmethod
    def _detect_columns(
        header_row: list[str],
        field_aliases: dict[str, set[str]],
    ) -> dict[str, int]:
        """Return {field_name: column_index} by matching header values against alias sets."""
        mapping: dict[str, int] = {}
        for field_name, aliases in field_aliases.items():
            for idx, header in enumerate(header_row):
                if header in aliases:
                    mapping[field_name] = idx
                    break
        return mapping

    @staticmethod
    def _raw_cell(row, col_map: dict[str, int], field: str):
        idx = col_map.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    @staticmethod
    def _str_cell(row, col_map: dict[str, int], field: str) -> str:
        val = ExcelParser._raw_cell(row, col_map, field)
        return str(val).strip() if val is not None else ""

    @staticmethod
    def _decimal_cell(row, col_map: dict[str, int], field: str) -> Decimal:
        val = ExcelParser._raw_cell(row, col_map, field)
        if val is None or val == "":
            return Decimal("0.00")
        try:
            return Decimal(str(val)).quantize(Decimal("0.01"))
        except InvalidOperation:
            return Decimal("0.00")

    @staticmethod
    def _parse_date(value) -> Optional[datetime]:
        """Convert a value to datetime.

        Handles:
          - Python datetime (openpyxl returns these for date-formatted cells)
          - Excel serial number (int/float) — days since 1899-12-30
          - String dates in common formats
        """
        if value is None:
            return None

        if isinstance(value, datetime):
            return value

        # Excel serial number (most common in this workbook)
        try:
            serial = int(float(str(value)))
            if 1 <= serial <= 2958465:  # valid Excel date range (1900-01-01 to 9999-12-31)
                return _EXCEL_EPOCH + timedelta(days=serial)
        except (ValueError, OverflowError):
            pass

        # String fallbacks
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d %b %Y"):
                try:
                    return datetime.strptime(value.strip(), fmt)
                except ValueError:
                    continue

        return None
